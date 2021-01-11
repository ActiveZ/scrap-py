import os
import requests
import pandas as pd
from citation import Citation
from bs4 import BeautifulSoup
# import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


os.system('clear')

r = requests.get("https://quotes.toscrape.com/")
soup = BeautifulSoup(r.content, "html.parser")
arr_citations = []

################################## CITATIONS ##################################################

print("-------- CITATIONS --------")
citations = soup.find_all("span", attrs={"class": "text", "itemprop": "text"})
for citation in citations:
    print(citation.text)
    citation = Citation(citation.text)
    arr_citations.append(citation)

################################## AUTHORS ##################################################

# encoding='utf-8'
print("-------- AUTHORS --------")


def create_xls(authorsSet):
    print("authorSet", authorsSet)
    wb = Workbook()
    dest_filename = 'authors_book.xlsx'
    ws1 = wb.active
    ws1.title = "authors"
    for row in range(1, 10):
    # for row in range(1, len(authorsSet)):
        ws1.append(range(2))

    # ws2 = wb.create_sheet(title="Pi")
    # ws2['F5'] = 3.14

    # wb.save(filename=dest_filename)
    df = pd.DataFrame([authorsSet]).T
    df.to_excel(excel_writer = "test.xlsx")

f = open("authors.txt", "w")
authors = soup.find_all(
    "small", attrs={"class": "author", "itemprop": "author"})
i = 0
authorsSet = []
for author in authors:
    arr_citations[i].author = author.text
    authorsSet.append(author.text)
    # print(author.text)
    f.write(author.text + "\n")
    i += 1
f.close()
authorsSet = list(set(authorsSet))
authorsSet.sort()
create_xls(authorsSet)


################################## TAGS ##################################################


print("-------- TAGS --------")
tags = soup.find_all(
    "meta", attrs={"class": "keywords", "itemprop": "keywords"})
# print(tags.text)


# remplissage d'un tableau de tags
arr_tags = []
for tag in tags:
    arr_tags.append(tag["content"])

i = 0
for tag in arr_tags:
    arr_citations[i].tags = tag
    i += 1

# élimination des doublons et tri alphabétique
tags = list(set(arr_tags))
tags.sort()
# print(tags)

# impression dans fichier tags.txt
f = open("tags.txt", "w")
for tag in tags:
    # print(tag)
    f.write(tag + "\n")
f.close()


################################## RESULTS ##################################################

f = open("quotes.md", "w")
for citation in arr_citations:
    f.write("citation:" + citation.content + " -- Auteur:" +
            citation.author + " -- Tags:" + citation.tags + "\n")
f.close()
